import streamlit as st
import openpyxl
import pandas as pd
from pydantic import ValidationError
from io import BytesIO
import datetime


class CSVCollector:
    def __init__(self, schema, aws):
        self._schema = schema
        self._aws = aws
        self._buffer = None
        self.cell_range = "C11"
        return
    
    def start(self):
        getData = self.getData()
        extractData = None
        validateData = None
        dfData = None
        if getData is not None:
            extractData = self.extractData(getData)
        if extractData is not None:
            validateData = self.validateData(extractData)
        if validateData is not None:
            dfData = self.transformDf(validateData)
        if dfData is not None:
            response = self.convert_to_parquer(dfData)
        if self._buffer is not None:
            file_name = self.file_name()
            self._aws.upload_file(response, file_name)
            return True
            
        return False
    
    def getData(self):
        dados_excel = st.file_uploader('Selecione um arquivo Excel', type=('xls', 'xlsx'))
        return dados_excel
    
    def extractData(self, dados_excel):
        workbook = openpyxl.load_workbook(dados_excel)
        sheet = workbook.active

        linha_inicial = 11

        # Contar quantas linhas existem a partir da linha 11
        linhas_preenchidas = sum(1 for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True) if any(row)) + 9
        self.cell_range = f"C11:I{linhas_preenchidas}"
        range_cell = sheet[self.cell_range]
        #pegando o meu indice 0, que é o cabeçalho
        headers = [cell.value for cell in range_cell[0]]

        data = []
        for row in range_cell[1:]:
            data.append([cell.value for cell in row])

        df = pd.DataFrame(data, columns=headers)
        return df

    # C:I12:I209
    def validateData(self, df):
        error = []
        valid_rows = []  # To store valid rows
        for index, row in df.iterrows():
            try:
                # Create an instance of the Pydantic model for each row
                row_dict = row.to_dict()
                valid_row = self._schema(**row.to_dict())
                #valid_rows.append(valid_row)  # Add the valid row to the list
            except ValidationError as e:
                # Append error message for rows that fail validation
                error.append(f"Erro na linha {index + 1}: {str(e)}")

        if error:
            st.error("\n".join(error))  # Displaying errors in Streamlit
            return None  # Return None if there are errors

        st.success("Arquivo Enviado com Sucesso!")
        return df
    
    def transformDf(self, response):
       result = pd.DataFrame(response)
       return result
    
    def convert_to_parquer(self, response):
        self._buffer = BytesIO()
        try:
            response.to_parquet(self._buffer)
            return self._buffer
        except:
            print('Erro ao converter para parquet')
            self._buffer = None

    def file_name(self):
        data_atual = datetime.datetime.now().isoformat()
        match = data_atual.split('.')
        return f"excel/excel-response-compras-{match[0]}.parquet"